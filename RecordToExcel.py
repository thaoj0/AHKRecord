# RecordToExcel.py by Jerry Thao
# This project uses tk Gui and openpyxl move files created by Record.ahk to Excel

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from tkinter import Label
from tkinter import StringVar
from tkinter import messagebox
import tkinter as tk
import os.path

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        #self.data = []
        #for i in range(5):
        #    self.data.append(tk.Entry(self))
        #    self.data[i].pack(side="left")
        self.label = Label(self, text="Database Path")
        self.label.pack(side="left")
        self.PathInString = StringVar()
        self.pathIn = tk.Entry(self,textvariable=self.PathInString)
        self.PathInString.set("Database/")
        self.pathIn.pack(side="left")

        self.label = Label(self, text="ExcelDatabase Path")
        self.label.pack(side="left")
        self.PathOutString = StringVar()
        self.pathOut = tk.Entry(self,textvariable=self.PathOutString)
        self.PathOutString.set("ExcelDatabase/")
        self.pathOut.pack(side="left")
        
        self.label = Label(self, text="Year (Double Digit)")
        self.label.pack(side="left")
        self.YearString = StringVar()
        self.year = tk.Entry(self,textvariable=self.YearString)
        self.YearString.set("0000")
        self.year.pack(side="left")

        self.label = Label(self, text="Month (Double Digit)")
        self.label.pack(side="left")
        self.MonthString = StringVar()
        self.month = tk.Entry(self,textvariable=self.MonthString)
        self.MonthString.set("00")
        self.month.pack(side="left")

        self.write = tk.Button(self, text="Transfer", command=self.loadTxtFile)
        self.write.pack(side="top")

        self.quit = tk.Button(self, text="QUIT", fg="red",command=self.master.destroy)
        self.quit.pack(side="bottom")

    def say_hi(self):
        print("hi there, everyone!")

    def openPy(self):
        workBook = Workbook()
        #loadBook = load_workbook("draw.xlsx")
        # grab the active worksheet
        workSheet = workBook.active
        #loadSheet = loadBook.active
        loadSheet = self.loadSheet("draw")

        # Python types will automatically be converted
        array = []
        for i,j in loadSheet.values:
            array.append([i,j])
        c = 0
        for i in self.data:
            c+=1
            array.append([i.get(),c])
        for i,j in array:
            workSheet.append([i,j])

        #import datetime
        #workSheet['A2'] = datetime.datetime.now()

        # Save the file
        #workBook.save("sample.xlsx")
        self.saveSheet(workBook,"sample")
    
    def loadSheet(self,fileString):
        filePath = fileString+".xlsx"
        if (os.path.isfile(filePath)):
            loadBook = load_workbook(filePath)
            loadSheet = loadBook.active
            return loadSheet
        else:
            print("File Not Found")
    
    def saveSheet(self,workBook,fileString):
        workBook.save(fileString+".xlsx")

    
    def loadTxtFile(self):
        workBook = Workbook()
        workSheet = workBook.active
        Months = ["","January", "February", "March", "April", "May", "June", 
        "July", "August", "September", "October", "November", "December"]
        workSheet.title = Months[int(self.MonthString.get())]
        workSheet.append(["Date","Cash","EBT","Credit","Taxable","Tax","Total","MP"])
        fileTo = self.PathOutString.get()+"Data"+self.YearString.get()+self.MonthString.get()+".xlsx"

        # Open each file that matches the month from days 1-31 and push to Excel
        # range is 33 because Excel arrays starts at 1 and not 0
        for i in range(1, 33):
            day = ""
            if(i<10):
                day ="0"+str(i)
            else:
                day = str(i)
            filePath = self.PathInString.get()+"Data"+self.YearString.get()+self.MonthString.get()+day+".txt"
            #print(filePath)
            if (os.path.isfile(filePath)):
                f = open(filePath)
                fList = f.read().split("{",1)[1].split("}",1)[0].split(",")
                f.close()
                arrayList = []
                dateIndex = True
                for x in fList:
                    #print( (x.strip("\n").split(":"))[1] )
                    value = (x.strip("\n").split(":"))[1] 
                    #If date becomes a float Excel will "Power" it automatically
                    #values won't =sum() if they are strings
                    if(dateIndex): 
                        dateIndex = False
                        value = "{}/{}/{}".format(value[0:4],value[4:6],value[6:8])
                    else:
                        value = float(value)
                    arrayList.append(value)
                workSheet.append(arrayList)
            else:
                print("File Not Found:"+filePath)
                workSheet.append(["","","","","","","",""]) #push in empty slot of missing dates
        workSheet.append(["Totals:","=Sum(B2:B31)","=Sum(C2:C31)","=Sum(D2:D31)","=Sum(E2:E31)","=Sum(F2:F31)","=Sum(G2:G31)","=Sum(H2:H31)"])

        red_fill = PatternFill(bgColor="FFC7CE")
        dxf = DifferentialStyle(fill=red_fill)
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
        workSheet.conditional_formatting.add("A1:C10", r)
        
        try:
            workBook.save(fileTo)
        except:
            messagebox.showerror("Task Failed","The Excel file must be closed:\n"+fileTo)
        else:
            messagebox.showinfo("Task Completed","New Excel File Created:\n"+fileTo)

root = tk.Tk()
app = Application(master=root)
app.master.minsize(100,25)
app.master.resizable(0,0)
app.mainloop()